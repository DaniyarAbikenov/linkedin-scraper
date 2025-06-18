import json
import os
import sys
from functools import reduce
from idlelib.editor import keynames
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller


# chromedriver_autoinstaller.install()  # Check if the current version of chromedriver exists
                                      # and if it doesn't exist, download it automatically,
                                      # then add chromedriver to path

def get_all_rows(worksheet, sheet_name):
    wb = load_workbook(filename=worksheet, data_only=True)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, [cell for cell in row])) for row in ws.iter_rows(min_row=2, values_only=True)]

def save_dict_to_excel(worksheet, data):
    wb = Workbook()
    ws = wb.active

    # Записываем заголовки (ключи из первого словаря)
    headers = list(data[0].keys())
    ws.append(headers)

    # Записываем данные
    for row in data:
        ws.append([row.setdefault(key, None) for key in headers])

    wb.save(worksheet)

def get_company_metadata(linkedin_link):
    options = webdriver.ChromeOptions()
    options.add_argument('--headless=new')
    driver = webdriver.Chrome(options=options)
    driver.get(linkedin_link)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        about_us = driver.find_element(By.CSS_SELECTOR, "[data-test-id='about-us']")
        table = about_us.find_element(By.CSS_SELECTOR, "dl.mt-6")
        dt_table = table.find_elements(By.CSS_SELECTOR, "dt") # Table key
        dd_table = table.find_elements(By.CSS_SELECTOR, "dd") # Table value
    except Exception as e:
        print("Error while checking URL " + linkedin_link)
        driver.close()
        return None
    data = dict(zip((dt.text for dt in dt_table), (dd.text for dd in dd_table)))
    driver.close()
    return data


def save_backup(data, file_name):
    backup_dir = Path("backups")
    os.makedirs(backup_dir, exist_ok=True)

    with open(backup_dir / (file_name + ".json"), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def raw_main(file_name):
    updated_data = []
    rows = get_all_rows(f"input/{file_name.replace(".xlsx", "")}.xlsx", "Sheet")
    try:
        for index, row in  enumerate(rows):
            print(f"{index + 1}. analyze {row['linkedin_link']}")
            company_info = get_company_metadata(row['linkedin_link'])
            if company_info:
                row.update(company_info)
                print(reduce(lambda a, b: f"{a} {b[0]}: {b[1]} |", row.items(), ""))
            updated_data.append(row)
    except Exception as e:
        print(e)
    finally:
        save_backup(updated_data, file_name)
        save_dict_to_excel(f"output/{file_name}.xlsx", updated_data)

def main():
    file_name = sys.argv[1] if len(sys.argv) > 1 else input("File name:")
    raw_main(file_name)


def multi_main():
    file_names = os.listdir('input')
    for file_name in file_names:
        print(file_name)
        raw_main(file_name)

if __name__ == "__main__":
    multi_main()